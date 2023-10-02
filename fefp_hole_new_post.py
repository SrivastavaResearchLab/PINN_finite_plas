import torch
import torch.nn as nn
import numpy as np
import random
import os
import scipy.io
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.interpolate import griddata
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable

torch.manual_seed(123)
np.random.seed(123)

class Model(nn.Module):
    def __init__(self, net_p, prob_p):
        super(Model, self).__init__()

        # Get parameters dictionaries
        self.net_p = net_p
        self.prob_p = prob_p

        # Break down the dictionaries
        neuron = self.net_p['neuron']
        lay_num = self.net_p['lay_num']
        b_mat = self.net_p['b_mat']
        var_in = 10 * self.net_p['var_in'] if b_mat else self.net_p['var_in']
        var_out = self.net_p['var_out']
        device = self.net_p['device']
        new_sp = self.net_p['new_sp']
        last_layer = self.net_p['last_layer']

        # B matrix to multiply sin/cos
        data_type = self.net_p['data_type']
        self.B = torch.tensor(np.random.randn(2, 10), dtype=data_type).to(device)
        self.C = torch.tensor(np.array([-5,-4,-3,-2,-1,1,2,3,4,5])[None, :], dtype=data_type).to(device)

        # Identity matrix (expand to batch dimension)
        self.iden = torch.eye(2, requires_grad=True, dtype=data_type)[None, :, :].to(device)

        # Create network structure
        # Add .double() after linear layer if using double precision
        self.net = nn.Sequential()
        self.net.add_module('Linear_layer_1', nn.Linear(var_in, neuron).double())
        self.net.add_module('Tanh_layer_1', nn.Tanh())
        for num in range(2, lay_num - 1):
            self.net.add_module('Linear_layer_%d' % num, nn.Linear(neuron, neuron).double())
            self.net.add_module('Tanh_layer_%d' % num, nn.Tanh())
        self.net.add_module('Linear_layer_%d' % (lay_num - 1), nn.Linear(neuron, neuron).double())

        if new_sp:
            # Beta parameter to control the softplus function
            self.beta = torch.tensor(5.0, requires_grad=True)
            self.last_layer = nn.Linear(neuron, var_out)
        else:
            if last_layer == 'Tanh':
                self.net.add_module('Tanh_layer_%d' % (lay_num - 1), nn.Tanh())
            elif last_layer == 'ReLU':
                self.net.add_module('ReLU_layer_%d' % (lay_num - 1), nn.ReLU())
            else:
                self.net.add_module('Tanh_layer_%d' % (lay_num - 1), nn.Softplus(beta=5))
            self.net.add_module('Linear_layer_out', nn.Linear(neuron, var_out).double())

    def forward(self, x):
        return self.net_norm(x)

    def softplus(self, r):
        return torch.log(1.0 + torch.exp(self.beta * r)) / self.beta

    def net_norm(self, r):
        # Normalize the input [x, y] into region [-1, 1]
        r = 2 * (r - 1/2)

        # Use B matrix if necessary
        cos = torch.cos(torch.matmul(r, self.B))
        sin = torch.sin(torch.matmul(r, self.B))

        # Use C matrix if necessary
        # cos = torch.matmul(r[:, 0:1], self.C)
        # sin = torch.matmul(r[:, 1:2], self.C)
        b_mat = self.net_p['b_mat']

        inp = torch.cat((cos, sin), dim=1) if b_mat else r
        temp = self.net(inp)
        if new_sp:
            return self.last_layer(self.softplus(temp))
        else:
            return temp

    def PK_to_cauchy(self, r):
        # Direct network output
        net_out = self.net_norm(r)
        ux = net_out[:, 0:1]
        uy = net_out[:, 1:2]
        P_xx = net_out[:, 2:3]
        P_yy = net_out[:, 3:4]
        P_xy = net_out[:, 4:5]
        P_yx = net_out[:, 5:6]

        P_new = torch.cat((torch.dstack((P_xx, P_xy)),
                              torch.dstack((P_yx, P_yy))), dim=1)

        # Deformation gradient in [2, 2] matrix
        ux_g = gradients(ux, r)[0]
        uy_g = gradients(uy, r)[0]
        FT_new = self.iden + torch.dstack((ux_g, uy_g))
        F_new = FT_new.transpose(1, 2)

        # Determinant of F
        det_F = torch.linalg.det(F_new)[:, None, None]

        # Cauchy stress
        T_new = torch.bmm(P_new, FT_new) / det_F
        return T_new


def gradients(outputs, inputs):
    return torch.autograd.grad(outputs, inputs, grad_outputs=torch.ones_like(outputs), create_graph=True)


def to_numpy(input):
    if isinstance(input, torch.Tensor):
        return input.detach().cpu().numpy()
    elif isinstance(input, np.ndarray):
        return input
    else:
        raise TypeError('Unknown type of input, expected torch.Tensor or '\
            'np.ndarray, but got {}'.format(type(input)))


def from_abaqus_xlsx(filepath):
    wb = load_workbook(filename=filepath)
    sheet = wb.active
    sheet_data = []
    for column in sheet.iter_cols(min_col=1,
                                  max_col=sheet.max_column,
                                  values_only=True):
        sheet_data.append(list(column))
    return np.array(sheet_data, dtype='float32')


def calculate_L2error():
    # Set the path to research drive
    os.chdir(path_drive)

    # Create list to store L2 absolute and relative errors
    L2_abs_u = []
    L2_abs_s = []
    L2_abs_peeq = []
    L2_rel_s = []
    L2_rel_u = []
    L2_rel_peeq = []

    ckp_num = int(max_epoch / ckp_div)
    # A loop to run trained model from research drive
    for i in range(ckp_num):
        model_ckp = 'ckp_' + str(ckp_div * (i + 1)) + '.pt'
        checkpoint = torch.load(model_ckp, map_location=torch.device('cpu'))
        model.load_state_dict(checkpoint['state_dict'])
        model.eval()

        # Reshape the PINN outputs
        net_out = to_numpy(model(xy_test))
        stress = to_numpy(model.PK_to_cauchy(xy_test))
        ux_pred = net_out[:, 0].reshape(num_test, num_test)
        uy_pred = net_out[:, 1].reshape(num_test, num_test)
        s_xx_pred = stress[:, 0, 0].reshape(num_test, num_test)
        s_yy_pred = stress[:, 1, 1].reshape(num_test, num_test)
        s_xy_pred = stress[:, 0, 1].reshape(num_test, num_test)
        peeq_pred = abs(net_out[:, 10]).reshape(num_test, num_test)

        # Depth-wise stack data
        u_pred = np.dstack((ux_pred, uy_pred))
        s_pred = np.dstack((s_xx_pred, s_yy_pred, 2*s_xy_pred))

        # Set values within the hole to be nan
        u_pred[flag, :] = np.nan
        s_pred[flag, :] = np.nan
        peeq_pred[flag] = np.nan

        # If this is the last epoch, call function to save the prediction
        if i == ckp_num - 1:
            # Set the working directory back to local
            os.chdir(path_local)
            model_pred = np.concatenate((u_pred, s_pred, peeq_pred[:, :, None]), axis=2)
            np.save('model_pred/' + model_name + '.npy', model_pred)

        # Remove nan values
        u_pred = u_pred[~np.isnan(u_pred)]
        s_pred = s_pred[~np.isnan(s_pred)]
        peeq_pred = peeq_pred[~np.isnan(peeq_pred)]

        # Calculate absolute error
        u_error = u_abaqus_ref - u_pred
        s_error = s_abaqus_ref - s_pred
        peeq_error = peeq_abaqus_ref - peeq_pred

        # L2 error for displacement, stress and peeq
        u_abs = np.sqrt(np.mean(u_error ** 2))
        L2_rel_u.append(u_abs / u_ref)
        L2_abs_u.append(u_abs)

        s_abs = np.sqrt(np.mean(s_error ** 2))
        L2_rel_s.append(s_abs / s_ref)
        L2_abs_s.append(s_abs)

        peeq_abs = np.sqrt(np.mean(peeq_error ** 2))
        L2_rel_peeq.append(peeq_abs / peeq_ref)
        L2_abs_peeq.append(peeq_abs)

    # Write L2 error to file
    L2_error = np.vstack((np.array(L2_rel_u), np.array(L2_rel_s), np.array(L2_rel_peeq)))
    np.save('L2_error/' + model_name + '.npy', L2_error)


"""Main code"""
# Select the model
model_name = 'fefp_hole_new_sp5_b_H05_BL_pow_coarser2_1234'

# Select what plot to use
niceplot = 1
lossplot = 0
indivplot = 0
L2plot = 0
errorplot = 0
pointerror = 0

# Network parameters
datatype = torch.float64
b_mat = True
new_sp = False
max_epoch = 1000000
ckp_div = 100000
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
if "relu" in model_name.lower():
    last_layer = 'ReLU'
elif "sp" in model_name.lower():
    last_layer = 'Softplus'
else:
    last_layer = 'Tanh'

net_p = {
    'var_in': 2,
    'var_out': 11,
    'lay_num': 7,
    'neuron': 30,
    'b_mat': b_mat,
    'data_type': datatype,
    'device': device,
    'last_layer': last_layer,
    'new_sp': new_sp
}

# Problem parameters
hole_r = 0.5
prob_p = {
    'E': 2,
    'nu': 0.3,
    'Y0': 0.4,
    'H0': 0.5,
    'npow': 0.7,
    'hole_r': hole_r
}

# Set paths
path_drive = 'H:/sniu3/Projects/PINN/Models/' + model_name
path_local = 'C:/Users/sniu3/Documents/python_work/PINN_2D'

# Test xy grid
num_test = 1001
x = np.linspace(0.0, 1.0, num=num_test)
y = np.linspace(0.0, 1.0, num=num_test)
y_grid, x_grid = np.meshgrid(y, x)
x_test = x_grid.flatten()[:, None]
y_test = y_grid.flatten()[:, None]
xy_test = np.hstack((x_test, y_test))
xy_test = torch.tensor(xy_test, requires_grad=True, dtype=datatype)
model = Model(net_p, prob_p)

# Load abaqus data
abaqus_data = from_abaqus_xlsx('abaqus_val_2D/Abaqus_pow_true.xlsx').T
x_abaqus = abaqus_data[:, 1:2]
y_abaqus = abaqus_data[:, 2:3]
coords_abaqus = np.hstack((x_abaqus, y_abaqus))
ux_abaqus = abaqus_data[:, 3:4]
uy_abaqus = abaqus_data[:, 4:5]
s_xx_abaqus = abaqus_data[:, 5:6]
s_yy_abaqus = abaqus_data[:, 6:7]
s_xy_abaqus = abaqus_data[:, 7:8]
peeq_abaqus = abaqus_data[:, 8:9]

# Get max and min values for plot
ux_min, ux_max = ux_abaqus.min(), ux_abaqus.max()
uy_min, uy_max = uy_abaqus.min(), uy_abaqus.max()
s_xx_min, s_xx_max = s_xx_abaqus.min(), s_xx_abaqus.max()
s_yy_min, s_yy_max = s_yy_abaqus.min(), s_yy_abaqus.max()
s_xy_min, s_xy_max = s_xy_abaqus.min(), s_xy_abaqus.max()
peeq_min, peeq_max = peeq_abaqus.min(), peeq_abaqus.max()

# Interpolation of abaqus data onto testing grids
ux_abaqus_int = griddata(coords_abaqus, ux_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)
uy_abaqus_int = griddata(coords_abaqus, uy_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)
peeq_abaqus_int = griddata(coords_abaqus, peeq_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)
s_xx_abaqus_int = griddata(coords_abaqus, s_xx_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)
s_yy_abaqus_int = griddata(coords_abaqus, s_yy_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)
s_xy_abaqus_int = griddata(coords_abaqus, s_xy_abaqus, (x_grid, y_grid)).reshape(num_test, num_test)

# Create flag to indicate all the indices within the hole
R = np.sqrt(x_grid ** 2 + y_grid ** 2)
flag = (R < 0.5)

# First manipulate Abaqus data to create reference L2 norm
u_abaqus_ref = np.dstack((ux_abaqus_int, uy_abaqus_int))
s_abaqus_ref = np.dstack((s_xx_abaqus_int, s_yy_abaqus_int, 2*s_xy_abaqus_int))
peeq_abaqus_ref = peeq_abaqus_int[:, :, None]

# Set values within the hole to be nan
u_abaqus_ref[flag, :] = np.nan
s_abaqus_ref[flag, :] = np.nan
peeq_abaqus_ref[flag, :] = np.nan

# Remove nan values
u_abaqus_ref = u_abaqus_ref[~np.isnan(u_abaqus_ref)]
s_abaqus_ref = s_abaqus_ref[~np.isnan(s_abaqus_ref)]
peeq_abaqus_ref = peeq_abaqus_ref[~np.isnan(peeq_abaqus_ref)]

# Reference L2 norm
u_ref = np.sqrt(np.mean(u_abaqus_ref ** 2))
s_ref = np.sqrt(np.mean(s_abaqus_ref ** 2))
peeq_ref = np.sqrt(np.mean(peeq_abaqus_ref ** 2))

try:
    L2_error = np.load('L2_error/' + model_name + '.npy')
except FileNotFoundError:
    calculate_L2error()
    L2_error = np.load('L2_error/' + model_name + '.npy')

# Unpack the L2 errors
L2_rel_u, L2_rel_s, L2_rel_peeq = L2_error[0, :], L2_error[1, :], L2_error[2, :]

# Get the prediction of model at the last epoch
prediction = np.load('model_pred/' + model_name + '.npy')
ux_pred, uy_pred = prediction[:, :, 0], prediction[:, :, 1]
s_xx_pred, s_yy_pred, s_xy_pred = prediction[:, :, 2], prediction[:, :, 3], prediction[:, :, 4]/2
peeq_pred = prediction[:, :, 5]

# Set values within the hole to be nan for plot purpose
ux_abaqus_int[flag] = np.nan
uy_abaqus_int[flag] = np.nan
s_xx_abaqus_int[flag] = np.nan
s_yy_abaqus_int[flag] = np.nan
s_xy_abaqus_int[flag] = np.nan
peeq_abaqus_int[flag] = np.nan

if niceplot:
    """ Plot for paper """
    uy_plot = [uy_pred, uy_abaqus_int]
    s_yy_plot = [s_yy_pred, s_yy_abaqus_int]
    peeq_plot = [peeq_pred, peeq_abaqus_int]
    for i in range(2):
        fig, ax = plt.subplots(1, 1, dpi=300)
        cs = ax.contourf(x_grid, y_grid, uy_plot[i], cmap='coolwarm',
                         levels=np.linspace(uy_min, uy_max, 100), vmin=uy_min, vmax=uy_max)
        divider = make_axes_locatable(ax)
        ax_cb = divider.new_horizontal(size="5%", pad=0.15)
        fig = ax.get_figure()
        fig.add_axes(ax_cb)
        ticks = np.linspace(uy_min, uy_max, 6)
        ticklabels = ['{:.3f}'.format(i) for i in ticks]
        cbar = fig.colorbar(cs, cax=ax_cb, ticks=ticks)
        cbar.ax.set_yticklabels(ticklabels, fontsize=16)
        ax_cb.yaxis.tick_right()
        ax.set_xticks([0.0, 0.5, 1.0])
        ax.set_yticks([0.0, 0.5, 1.0])
        ax.set_aspect('equal', 'box')
        ax.set_xticklabels([0.0, 0.5, 1.0], fontsize=16)
        ax.set_yticklabels([0.0, 0.5, 1.0], fontsize=16)
        fig.show()

    for i in range(2):
        fig, ax = plt.subplots(1, 1, dpi=300)
        cs = ax.contourf(x_grid, y_grid, s_yy_plot[i], cmap='coolwarm',
                         levels=np.linspace(s_yy_min, s_yy_max, 100), vmin=s_yy_min, vmax=s_yy_max)
        divider = make_axes_locatable(ax)
        ax_cb = divider.new_horizontal(size="5%", pad=0.15)
        fig = ax.get_figure()
        fig.add_axes(ax_cb)
        ticks = np.linspace(s_yy_min, s_yy_max, 6)
        ticklabels = ['{:.3f}'.format(i) for i in ticks]
        cbar = fig.colorbar(cs, cax=ax_cb, ticks=ticks)
        cbar.ax.set_yticklabels(ticklabels)
        ax_cb.yaxis.tick_right()
        ax.set_xticks([0.0, 0.5, 1.0])
        ax.set_yticks([0.0, 0.5, 1.0])
        ax.set_aspect('equal', 'box')
        fig.show()

    for i in range(2):
        fig, ax = plt.subplots(1, 1, dpi=300)
        cs = ax.contourf(x_grid, y_grid, peeq_plot[i], cmap='coolwarm',
                         levels=np.linspace(peeq_min, peeq_max, 100), vmin=peeq_min, vmax=peeq_max)
        divider = make_axes_locatable(ax)
        ax_cb = divider.new_horizontal(size="5%", pad=0.15)
        fig = ax.get_figure()
        fig.add_axes(ax_cb)
        ticks = np.linspace(peeq_min, peeq_max, 6)
        ticklabels = ['{:.3f}'.format(i) for i in ticks]
        cbar = fig.colorbar(cs, cax=ax_cb, ticks=ticks)
        cbar.ax.set_yticklabels(ticklabels)
        ax_cb.yaxis.tick_right()
        ax.set_xticks([0.0, 0.5, 1.0])
        ax.set_yticks([0.0, 0.5, 1.0])
        ax.set_aspect('equal', 'box')
        fig.show()

    '''
    # Figure for displacement
    fig, axes = plt.subplots(2, 2, figsize=(10, 8))
    ax1 = axes[0, 0]
    im1 = ax1.contourf(x_grid, y_grid, ux_pred, 100, cmap='coolwarm', vmin=ux_min, vmax=ux_max)
    ax1.set_aspect(1)
    ax1.set_title('$u_x$ PINN')
    ax2 = axes[0, 1]
    im2 = ax2.contourf(x_grid, y_grid,  ux_abaqus_int, 100, cmap='coolwarm', vmin=ux_min, vmax=ux_max)
    fig.colorbar(im2, ax=ax2)
    ax2.set_aspect(1)
    ax2.set_title('$u_x$ Abaqus')
    ax3 = axes[1, 0]
    im3 = ax3.contourf(x_grid, y_grid, uy_pred, 100, cmap='coolwarm', vmin=uy_min, vmax=uy_max)
    ax3.set_aspect(1)
    ax3.set_title('$u_y$ PINN')
    ax4 = axes[1, 1]
    im4 = ax4.contourf(x_grid, y_grid, uy_abaqus_int, 100, cmap='coolwarm', vmin=uy_min, vmax=uy_max)
    fig.colorbar(im4, ax=ax4)
    ax4.set_aspect(1)
    ax4.set_title('$u_y$ Abaqus')
    fig.show()

    # Figure for stress
    fig, axes = plt.subplots(3, 2, figsize=(8, 10))
    ax1 = axes[0, 0]
    im1 = ax1.contourf(x_grid, y_grid, s_xx_pred, 100, cmap='coolwarm', vmin=s_xx_min, vmax=s_xx_max)
    ax1.set_aspect(1)
    ax1.set_title('$\sigma_{xx}$ PINN')
    ax2 = axes[0, 1]
    im2 = ax2.contourf(x_grid, y_grid,  s_xx_abaqus_int, 100, cmap='coolwarm', vmin=s_xx_min, vmax=s_xx_max)
    fig.colorbar(im2, ax=ax2)
    ax2.set_aspect(1)
    ax2.set_title('$\sigma_{xx}$ Abaqus')
    ax3 = axes[1, 0]
    im3 = ax3.contourf(x_grid, y_grid, s_yy_pred, 100, cmap='coolwarm', vmin=s_yy_min, vmax=s_yy_max)
    ax3.set_aspect(1)
    ax3.set_title('$\sigma_{yy}$ PINN')
    ax4 = axes[1, 1]
    im4 = ax4.contourf(x_grid, y_grid, s_yy_abaqus_int, 100, cmap='coolwarm', vmin=s_yy_min, vmax=s_yy_max)
    fig.colorbar(im4, ax=ax4)
    ax4.set_aspect(1)
    ax4.set_title('$\sigma_{yy}$ Abaqus')
    ax5 = axes[2, 0]
    im5 = ax5.contourf(x_grid, y_grid, s_xy_pred, 100, cmap='coolwarm', vmin=s_xy_min, vmax=s_xy_max)
    ax5.set_aspect(1)
    ax5.set_title('$\sigma_{xy}$ PINN')
    ax6 = axes[2, 1]
    im6 = ax6.contourf(x_grid, y_grid, s_xy_abaqus_int, 100, cmap='coolwarm', vmin=s_xy_min, vmax=s_xy_max)
    fig.colorbar(im6, ax=ax6)
    ax6.set_aspect(1)
    ax6.set_title('$\sigma_{xy}$ Abaqus')
    plt.show()

    # Figure for plastic strain
    fig, axes = plt.subplots(1, 2, figsize=(8.6, 4))
    ax1 = axes[0]
    im1 = ax1.contourf(x_grid, y_grid, peeq_pred, 100, cmap='coolwarm', vmin=peeq_min, vmax=peeq_max)
    ax1.set_title('$\epsilon^p$ PINN')
    ax2 = axes[1]
    im2 = ax2.contourf(x_grid, y_grid,  peeq_abaqus_int, 100, cmap='coolwarm', vmin=peeq_min, vmax=peeq_max)
    ax2.set_title('$\epsilon^p$ Abaqus')
    fig.colorbar(im2, fraction=0.05, pad=0.04, format='%.2f')
    fig.show() '''

""" Plot loss """
if lossplot:
    # Make sure plot does not exceed limit
    mpl.rcParams['agg.path.chunksize'] = 10000

    # Go to research drive to load the loss
    os.chdir(path_drive)
    loss = np.load('loss.npy')
    os.chdir(path_local)
    epoch = np.linspace(0, 1000000, num=10000)
    loss_total, loss_pde, loss_bc, loss_const = loss[0, 0:-1:100], loss[1, 0:-1:100], loss[2, 0:-1:100], loss[3, 0:-1:100]
    fig, ax = plt.subplots(1, 1)
    ax.plot(loss_pde)
    ax.plot(loss_bc)
    ax.plot(loss_const)
    ax.legend(['PDE', 'BC', 'Constitutive'])
    ax.set_xlabel('Epochs')
    ax.set_yscale('log')
    ax.set_title('Total loss')
    scipy.io.savemat('loss_coarser.mat', dict(epoch=epoch, loss_total=loss_total, loss_pde=loss_pde,
                                      loss_bc=loss_bc, loss_const=loss_const))
    plt.show()

""" Plot L2 error evolution """
if L2plot:
    # mpl.rcParams.update({'font.size': 35})
    # plt.rcParams["font.family"] = "Times New Roman"
    # fig, ax = plt.subplots(1, 1, figsize=(16, 10), dpi=150)
    # ax.plot(epoch, L2_rel_u, '--ok', markersize=20)
    # ax.plot(epoch, L2_rel_s, '--^k', markersize=20)
    # ax.plot(epoch, L2_rel_peeq, '--sk', markersize=20)
    epoch = [ckp_div * (i+1) for i in range(int(max_epoch/ckp_div))]
    fig, ax = plt.subplots(1, 1)
    ax.plot(epoch, L2_rel_u, '--ok')
    ax.plot(epoch, L2_rel_s, '--^k')
    ax.plot(epoch, L2_rel_peeq, '--sk')
    ax.legend(['Displacement', 'Stress', 'Plastic Strain'])
    ax.set_xlabel('Epochs')
    ax.set_yscale('log')
    ax.set_title('L2 relative error')
    fig.show()
    print(L2_rel_u[-3:].mean(), L2_rel_s[-3:].mean(), L2_rel_peeq[-3:].mean())

if errorplot:
    fig, ax = plt.subplots(1, 1, dpi=300)
    z = peeq_abaqus_int - peeq_pred
    z = z[~np.isnan(z)]
    cs = ax.contourf(x_grid, y_grid, peeq_abaqus_int - peeq_pred, cmap='coolwarm', levels=np.linspace(np.min(z), np.max(z), 100), vmin=-0.004, vmax=0.004)
    divider = make_axes_locatable(ax)
    ax_cb = divider.new_horizontal(size="5%", pad=0.15)
    fig = ax.get_figure()
    fig.add_axes(ax_cb)
    ticks = np.linspace(np.min(z), np.max(z), 6)
    ticklabels = ['{:.3f}'.format(i) for i in ticks]
    cbar = fig.colorbar(cs, cax=ax_cb, ticks=ticks)
    cbar.ax.set_yticklabels(ticklabels)
    ax_cb.yaxis.tick_right()
    ax.set_xticks([0.0, 0.5, 1.0])
    ax.set_yticks([0.0, 0.5, 1.0])
    ax.set_aspect('equal', 'box')
    # plt.savefig('nob.eps', format='eps',dpi=300)
    # im = ax.contourf(x_grid, y_grid, peeq_abaqus_int - peeq_pred, 100, cmap='coolwarm', vmin=-0.004, vmax=0.004)
    # ax.set_aspect(1)
    # ax.set_title('$\epsilon^p$ pointwise error')
    # fig.colorbar(im, ax=ax)
    fig.show()

if pointerror:
    os.chdir(path_drive)
    model = torch.load('fefp_plas_hole.pt', map_location=torch.device('cpu'))
    model.eval()

    # Specify points
    pt_coord = np.array([[0.5, 0], [0.6, 0], [0.9, 0]])
    pt_coord = torch.tensor(pt_coord, requires_grad=True, dtype=datatype)
    net_out = to_numpy(model(pt_coord))
    stress = to_numpy(model.PK_to_cauchy(pt_coord))
    s_yy_abaqus_pt = np.array([s_yy_abaqus_int[500, 0], s_yy_abaqus_int[600, 0], s_yy_abaqus_int[900, 0]])
    peeq_abaqus_pt = np.array([peeq_abaqus_int[500, 0], peeq_abaqus_int[600, 0], peeq_abaqus_int[900, 0]])
    s_yy_pt_error = abs(stress[:, 1, 1] - s_yy_abaqus_pt)
    peeq_pt_error = abs(abs(net_out[:, 10])- peeq_abaqus_pt)
    os.chdir(path_local)
    print(s_yy_pt_error, peeq_pt_error)


""" Plot the plastic strain in the bottom boundary as in 1D """
fig, ax = plt.subplots(1, 1)
ax.plot(peeq_pred[:, 0])
ax.plot(peeq_abaqus_int[:, 0])
ax.legend(['PINN', 'Abaqus'])
ax.set_title('Plastic strain in the bottom symmetry boundary')
fig.show()

# Print the L2 error
print("----------------------------------------------------------")
print("For model: " + model_name.upper())
print("L2 relative error of DISPLACEMENT: %.5f" % (L2_rel_u[-1]*100), "%")
print("L2 relative error of STRESS: %.5f" % (L2_rel_s[-1]*100), "%")
print("L2 relative error of PLASTIC STRAIN: %.5f" % (L2_rel_peeq[-1]*100), "%")
print("----------------------------------------------------------")
print("Baseline (NEW_SP5_B_H05_BL_POW) results:")
print("L2 relative error of DISPLACEMENT: 0.02287 %")
print("L2 relative error of STRESS: 0.21132 %")
print("L2 relative error of PLASTIC STRAIN: 0.64033 %")
print("----------------------------------------------------------")