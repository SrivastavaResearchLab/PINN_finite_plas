"""
Author: Sijun Niu
    PINNs (physical-informed neural network) for solving 2D tension on a plate with a hole
    Large deformation with elasto-plastic material
"""

import sys
# sys.path.insert(0,'../../Utils')

import torch
import torch.nn as nn
import numpy as np
import time
import random
import math
import matplotlib as mpl
from openpyxl import load_workbook
import matplotlib.pyplot as plt

torch.manual_seed(123)
np.random.seed(123)

class Model(nn.Module):
    def __init__(self, net_p, prob_p):
        super(Model, self).__init__()

        # Get parameters dictionaries
        self.net_p = net_p
        self.prob_p = prob_p

        # Break down the dictionaries
        neuron = self.net_p['neuron']
        lay_num = self.net_p['lay_num']
        b_mat = self.net_p['b_mat']
        var_in = 10 * self.net_p['var_in'] if b_mat else self.net_p['var_in']
        var_out = self.net_p['var_out']
        device = self.net_p['device']

        # Counting the epochs
        self.current_epoch = 0

        # B matrix to multiply sin/cos
        data_type = self.net_p['data_type']
        self.B = torch.tensor(np.random.randn(2, 10), dtype=data_type).to(device)
        self.C = torch.tensor(np.array([-5,-4,-3,-2,-1,1,2,3,4,5])[None, :], dtype=data_type).to(device)

        # Identity matrix (expand to batch dimension)
        self.iden = torch.eye(2, requires_grad=True, dtype=data_type)[None, :, :].to(device)

        # Create network structure
        # Add .double() after linear layer if using double precision
        self.net = nn.Sequential()
        self.net.add_module('Linear_layer_1', nn.Linear(var_in, neuron).double())
        self.net.add_module('Tanh_layer_1', nn.Tanh())
        for num in range(2, lay_num - 1):
            self.net.add_module('Linear_layer_%d' % num, nn.Linear(neuron, neuron).double())
            self.net.add_module('Tanh_layer_%d' % num, nn.Tanh())
        self.net.add_module('Linear_layer_%d' % (lay_num - 1), nn.Linear(neuron, neuron).double())
        self.net.add_module('ReLU_layer_%d' % (lay_num - 1), nn.Softplus(beta=5))
        self.net.add_module('Linear_layer_out', nn.Linear(neuron, var_out).double())

    def forward(self, x):
        return self.net_norm(x)

    def update_epoch(self, epoch):
        self.current_epoch = epoch

    def net_norm(self, r):
        # Normalize the input [x, y] into region [-1, 1]
        r = 2 * (r - 1/2)

        b_mat = self.net_p['b_mat']
        # Use B matrix if necessary
        cos = torch.cos(torch.matmul(r, self.B))
        sin = torch.sin(torch.matmul(r, self.B))

        inp = torch.cat((cos, sin), dim=1) if b_mat else r
        return self.net(inp)

    def loss_pde(self, r):
        net_out = self.net_norm(r)
        P_xx = net_out[:, 2:3]
        P_yy = net_out[:, 3:4]
        P_xy = net_out[:, 4:5]
        P_yx = net_out[:, 5:6]

        P_xx_x = gradients(P_xx, r)[0][:, 0:1]
        P_yy_y = gradients(P_yy, r)[0][:, 1:2]
        P_xy_y = gradients(P_xy, r)[0][:, 1:2]
        P_yx_x = gradients(P_yx, r)[0][:, 0:1]

        loss = ((P_xx_x + P_xy_y) ** 2).mean() + ((P_yx_x + P_yy_y) ** 2).mean()
        return loss

    def loss_bc(self, bc_l_train, bc_r_train, bc_b_train, bc_t_train, bc_hole_train):
        # Left BC: ux = 0, P_yx = 0
        net_out_l = self.net_norm(bc_l_train)
        ux_l = net_out_l[:, 0:1]
        P_yx_l = net_out_l[:, 5:6]
        loss_l = (ux_l**2).mean() + (P_yx_l**2).mean()

        # Right BC: P_xx = 0, P_yx = 0
        net_out_r = self.net_norm(bc_r_train)
        P_xx_r = net_out_r[:, 2:3]
        P_yx_r = net_out_r[:, 5:6]
        loss_r = (P_xx_r**2).mean() + (P_yx_r**2).mean()

        # Bottom BC: uy = 0, P_xy = 0
        net_out_b = self.net_norm(bc_b_train)
        uy_b = net_out_b[:, 1:2]
        P_xy_b = net_out_b[:, 4:5]
        loss_b = (uy_b**2).mean() + (P_xy_b**2).mean()

        # Top BC: P_yy = p, P_xy = 0
        net_out_t = self.net_norm(bc_t_train)
        P_yy_t = net_out_t[:, 3:4]
        P_xy_t = net_out_t[:, 4:5]
        p = 0.18
        loss_xy_t = P_xy_t
        loss_yy_t = P_yy_t - p
        loss_t = (loss_xy_t**2).mean() + (loss_yy_t**2).mean()

        # Hole BC: traction free
        net_out_hole = self.net_norm(bc_hole_train)
        P_xx_hole = net_out_hole[:, 2:3]
        P_yy_hole = net_out_hole[:, 3:4]
        P_xy_hole = net_out_hole[:, 4:5]
        P_yx_hole = net_out_hole[:, 5:6]
        loss_hole_1 = P_xx_hole * bc_hole_train[:, 0:1] + P_xy_hole * bc_hole_train[:, 1:2]
        loss_hole_2 = P_yx_hole * bc_hole_train[:, 0:1] + P_yy_hole * bc_hole_train[:, 1:2]
        loss_hole = (loss_hole_1 ** 2).mean() + (loss_hole_2 ** 2).mean()

        return loss_l + loss_r + loss_b + loss_t + loss_hole

    def loss_const(self, r):
        # Direct network output
        net_out = self.net_norm(r)
        ux = net_out[:, 0:1]
        uy = net_out[:, 1:2]
        P_xx = net_out[:, 2:3]
        P_yy = net_out[:, 3:4]
        P_xy = net_out[:, 4:5]
        P_yx = net_out[:, 5:6]
        Fp_new_xx = net_out[:, 6:7] + 1
        Fp_new_xy = net_out[:, 7:8]
        Fp_new_yx = net_out[:, 8:9]
        Fp_new_yy = net_out[:, 9:10] + 1
        del_eps_p_bar = abs(net_out[:, 10:11])

        # Parameters
        E = self.prob_p['E']
        nu = self.prob_p['nu']
        Y0 = self.prob_p['Y0']
        H0 = self.prob_p['H0']
        npow = self.prob_p['npow']
        K = E / 3 / (1 - 2 * nu)
        G = E / 2 / (1 + nu)

        # Plastic deformation gradient components (1 diagonal and 0 off-diagonal)
        Fp_xx = torch.ones_like(Fp_new_xx)
        Fp_xy = torch.zeros_like(Fp_new_xy)
        Fp_yx = torch.zeros_like(Fp_new_yx)
        Fp_yy = torch.ones_like(Fp_new_yy)

        # Initialize eps_p_bar_new and eps_p_bar
        eps_p_bar = torch.zeros_like(del_eps_p_bar)
        eps_p_bar_new = del_eps_p_bar

        # Assemble the plastic deformation gradient at t=n and t=n+1
        Fp = torch.concat((torch.dstack((Fp_xx, Fp_xy)),
                           torch.dstack((Fp_yx, Fp_yy))), dim=1)

        # Yield surface
        Y = Y0 + H0 * eps_p_bar ** npow
        Y_new = Y0 + H0 * eps_p_bar_new ** npow

        # Trial deformation gradient in [2, 2] matrix
        ux_g = gradients(ux, r)[0]
        uy_g = gradients(uy, r)[0]
        FT_tr = self.iden + torch.dstack((ux_g, uy_g))
        F_tr = FT_tr.transpose(1, 2)

        # Right Cauchy Green tensor
        C_tr = torch.bmm(FT_tr, F_tr)

        # Eigenvalues: C->U->E
        c_eig, Q = torch.linalg.eigh(C_tr)
        u_eig = c_eig ** 0.5
        u_inv_eig = 1 / u_eig
        e_eig = torch.log(u_eig)
        e_diag = torch.diag_embed(e_eig)

        # Calculate Rotation tensor
        u_inv_diag = torch.diag_embed(u_inv_eig)
        QT = Q.transpose(1, 2)
        U_tr_inv = torch.bmm(Q, torch.bmm(u_inv_diag, QT))
        R_tr = torch.bmm(F_tr, U_tr_inv)

        # Hencky strain (and its deviatoric part)
        E_tr = torch.bmm(Q, torch.bmm(e_diag, QT))
        tr_E_tr = (E_tr[:, 0, 0] + E_tr[:, 1, 1])[:, None, None]
        E0_tr = E_tr - tr_E_tr / 3 * self.iden

        # Mandel stress in terms of the Hencky strain (and its deviatoric part)
        M_tr = 2 * G * E0_tr + K * tr_E_tr * self.iden

        # Additional z direction Mandel stress
        M_tr_xx = M_tr[:, 0:1, 0:1]
        M_tr_yy = M_tr[:, 1:2, 1:2]
        M_tr_xy = M_tr[:, 0:1, 1:2]
        M_tr_zz = nu * (M_tr_xx + M_tr_yy)

        tr_M_tr = (M_tr_xx + M_tr_yy + M_tr_zz)
        M0_tr = M_tr - tr_M_tr / 3 * self.iden

        # Calculate Mises stress
        sig_bar_tr = (((M_tr_xx - M_tr_yy) ** 2 +
                       (M_tr_yy - M_tr_zz) ** 2 +
                       (M_tr_zz - M_tr_xx) ** 2) / 2 +
                      3 * M_tr_xy ** 2) ** 0.5
        f_tr = (sig_bar_tr - Y[:, :, None]).squeeze(-1)

        # Plastic direction
        Np_tr = M0_tr / sig_bar_tr * 1.5 ** 0.5

        # (Non)-linear plastic equation g(eps_bar_p)
        plas_eq = sig_bar_tr.squeeze(-1) - 3 * G * del_eps_p_bar - Y_new

        # Update the Mandel stress
        M_new = M_tr - 6 ** 0.5 * G * del_eps_p_bar[:, :, None] * Np_tr

        # Convert the Mandel stress to 1st PK stress
        RT_tr = R_tr.transpose(1, 2)
        Pe_new = torch.bmm(R_tr, torch.bmm(M_new, RT_tr))
        PeT_new = Pe_new.transpose(1, 2)
        PT_new = torch.linalg.solve(F_tr, PeT_new)
        P_new = PT_new.transpose(1, 2)

        # Update the plastic stretch
        Dp = 1.5 ** 0.5 * del_eps_p_bar[:, :, None] * Np_tr
        dp_eig, V = torch.linalg.eigh(Dp)
        dp_exp_eig = torch.exp(dp_eig)
        dp_exp_diag = torch.diag_embed(dp_exp_eig)
        VT = V.transpose(1, 2)
        Dp_exp = torch.bmm(V, torch.bmm(dp_exp_diag, VT))

        # Another plastic deformation gradient from equivalent plastic strain
        Fp_new_fromD = torch.bmm(Dp_exp, Fp)

        # Plastic equation loss
        loss_plas = (f_tr < 0) * del_eps_p_bar + (f_tr >= 0) * plas_eq

        # Stress loss
        loss_P_xx = P_new[:, 0:1, 0] - P_xx
        loss_P_yy = P_new[:, 1:2, 1] - P_yy
        loss_P_xy = P_new[:, 0:1, 1] - P_xy
        loss_P_yx = P_new[:, 1:2, 0] - P_yx

        # Plastic deformation gradient loss
        loss_Fp_xx = Fp_new_fromD[:, 0:1, 0] - Fp_new_xx
        loss_Fp_yy = Fp_new_fromD[:, 1:2, 1] - Fp_new_yy
        loss_Fp_xy = Fp_new_fromD[:, 0:1, 1] - Fp_new_xy
        loss_Fp_yx = Fp_new_fromD[:, 1:2, 0] - Fp_new_yx

        return (loss_plas ** 2).mean() + \
               (loss_Fp_xx ** 2).mean() + (loss_Fp_yy ** 2).mean() + \
               (loss_Fp_xy ** 2).mean() + (loss_Fp_yx ** 2).mean() + \
               (loss_P_xx ** 2).mean() + (loss_P_yy ** 2).mean() + \
               (loss_P_xy ** 2).mean() + (loss_P_yx ** 2).mean()

    def PK_to_cauchy(self, r):
        # Direct network output
        net_out = self.net_norm(r)
        ux = net_out[:, 0:1]
        uy = net_out[:, 1:2]
        P_xx = net_out[:, 2:3]
        P_yy = net_out[:, 3:4]
        P_xy = net_out[:, 4:5]
        P_yx = net_out[:, 5:6]

        P_new = torch.concat((torch.dstack((P_xx, P_xy)),
                              torch.dstack((P_yx, P_yy))), dim=1)

        # Deformation gradient in [2, 2] matrix
        ux_g = gradients(ux, r)[0]
        uy_g = gradients(uy, r)[0]
        FT_new = self.iden + torch.dstack((ux_g, uy_g))
        F_new = FT_new.transpose(1, 2)

        # Determinant of F
        det_F = torch.linalg.det(F_new)[:, None, None]

        # Cauchy stress
        T_new = torch.bmm(P_new, FT_new) / det_F
        return T_new

def gradients(outputs, inputs):
    return torch.autograd.grad(outputs, inputs, grad_outputs=torch.ones_like(outputs), create_graph=True)


def to_numpy(input):
    if isinstance(input, torch.Tensor):
        return input.detach().cpu().numpy()
    elif isinstance(input, np.ndarray):
        return input
    else:
        raise TypeError('Unknown type of input, expected torch.Tensor or '\
            'np.ndarray, but got {}'.format(type(input)))


def from_abaqus_xlsx(filepath):
    wb = load_workbook(filename=filepath)
    sheet = wb.active
    sheet_data = []
    for column in sheet.iter_cols(min_col=1,
                                  max_col=sheet.max_column,
                                  values_only=True):
        sheet_data.append(list(column))
    return np.array(sheet_data, dtype='float32')


def save_loss(total, pde, bc, const):
    total = np.array(total)
    pde = np.array(pde)
    bc = np.array(bc)
    const = np.array(const)
    loss = np.vstack((total, pde, bc, const))
    np.save('loss.npy', loss)

    # Plot the loss every time to make sure that it did not explode
    fig, ax = plt.subplots(1, 1)
    ax.plot(total)
    ax.set_xlabel('Epochs')
    ax.set_yscale('log')
    ax.set_title('Total loss')
    plt.savefig('loss.png', dpi=150)


def main():
    # Make sure plot does not exceed limit
    mpl.rcParams['agg.path.chunksize'] = 10000

    # Network parameters
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(torch.cuda.is_available())
    epochs = 3000000
    lr = 0.001
    datatype = torch.float64
    decay_rate = 0.5 ** (1 / 1000000)
    b_mat = True

    net_p = {
        'var_in': 2,
        'var_out': 11,
        'lay_num': 5,
        'neuron': 30,
        'b_mat': b_mat,
        'data_type': datatype,
        'device': device
    }

    # Training data parameters
    num_train = 21
    num_hole = 15

    # Problem parameters
    hole_r = 0.5
    prob_p = {
        'E': 2,
        'nu': 0.3,
        'Y0': 0.4,
        'H0': 0.5,
        'npow': 0.7,
        'hole_r': hole_r
    }

    node_abaqus = from_abaqus_xlsx('Abaqus_fine_mesh.xlsx').T
    bc_l_list = [2,  3,  5, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44,
                 45, 46, 47, 48, 49]
    bc_t_list = [1,   2,  11,  12,  13,  14,  15,  16,  17,  18,  19,  20,  21,  22,  23,  24,
                 25,  26,  27,  28,  29,  30,  31, 222, 223, 224, 225, 226, 227, 228, 229, 230,
                 231, 232, 233, 234, 235, 236, 237, 238, 239]
    bc_r_list = [9,  10,  11, 166, 167, 168, 169, 170, 171, 172, 173, 174, 175, 176, 177, 178,
                 179, 180, 181, 182, 183, 184, 185, 204, 205, 206, 207, 208, 209, 210, 211, 212,
                 213, 214, 215, 216, 217, 218, 219, 220, 221]
    bc_b_list = [7,   8,   9, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157, 158, 159, 160,
                 161, 162, 163, 164, 165]
    bc_hole_list = [5,   6,   7,  88,  89,  90,  91,  92,  93,  94,  95,  96,  97,  98,  99, 100,
                    101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116,
                    117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127]
    bc_l = np.array(bc_l_list) - 1
    bc_t = np.array(bc_t_list) - 1
    bc_r = np.array(bc_r_list) - 1
    bc_b = np.array(bc_b_list) - 1
    bc_hole = np.array(bc_hole_list) - 1
    bc_l_train = node_abaqus[bc_l, 1:3]
    bc_t_train = node_abaqus[bc_t, 1:3]
    bc_r_train = node_abaqus[bc_r, 1:3]
    bc_b_train = node_abaqus[bc_b, 1:3]
    bc_hole_train = node_abaqus[bc_hole, 1:3]
    xy_train = node_abaqus[:, 1:3]

    # Make the numpy arrays to tensor
    bc_l_train = torch.tensor(bc_l_train, requires_grad=True, dtype=datatype).to(device)
    bc_r_train = torch.tensor(bc_r_train, requires_grad=True, dtype=datatype).to(device)
    bc_b_train = torch.tensor(bc_b_train, requires_grad=True, dtype=datatype).to(device)
    bc_t_train = torch.tensor(bc_t_train, requires_grad=True, dtype=datatype).to(device)
    bc_hole_train = torch.tensor(bc_hole_train, requires_grad=True, dtype=datatype).to(device)
    xy_train = torch.tensor(xy_train, requires_grad=True, dtype=datatype).to(device)

    # Create model
    model = Model(net_p, prob_p).to(device)
    print(model)

    # Loss and optimizer
    optimizer = torch.optim.Adam(model.parameters(), lr=lr)
    scheduler = torch.optim.lr_scheduler.ExponentialLR(optimizer, decay_rate)

    def closure():
        optimizer.zero_grad()
        loss_pde = model.loss_pde(xy_train)
        loss_bc = model.loss_bc(bc_l_train, bc_r_train, bc_b_train, bc_t_train, bc_hole_train)
        loss_const = model.loss_const(xy_train)
        loss = loss_pde + loss_bc + loss_const
        loss.backward()
        return loss, loss_pde, loss_bc, loss_const

    # Start training
    print('start training...')
    loss_tot = []
    loss_pde = []
    loss_bc = []
    loss_const = []
    tic = time.time() / 3600

    for epoch in range(epochs):
        model.train()
        model.update_epoch(epoch)
        loss_i, loss_pde_i, loss_bc_i, loss_const_i = optimizer.step(closure)
        scheduler.step()
        loss_tot.append(loss_i.item())
        loss_pde.append(loss_pde_i.item())
        loss_bc.append(loss_bc_i.item())
        loss_const.append(loss_const_i.item())

        if (epoch + 1) % 100000 == 0:
            checkpoint = {
                'epoch': epoch + 1,
                'state_dict': model.state_dict(),
                'optimizer': optimizer.state_dict()
            }
            f_path = 'ckp_' + str(epoch + 1) + '.pt'
            torch.save(checkpoint, f_path)

            # Save the loss and plot each time
            save_loss(loss_tot, loss_pde, loss_bc, loss_const)

    # Display training time
    toc = time.time() / 3600
    print(f'total training time: {toc-tic}')

    # Save the model and write the model parameters to file
    model_path = "fefp_plas_hole.pt"
    torch.save(model, model_path)

if __name__ == '__main__':
    main()
